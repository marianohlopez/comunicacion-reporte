from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import os
from dotenv import load_dotenv
import yagmail

load_dotenv()

MAIL_AUTOR = os.getenv("MAIL_AUTOR")
APP_GMAIL_PASS = os.getenv("APP_GMAIL_PASS")
MAIL_DESTINO = os.getenv("MAIL_DESTINO")

today = datetime.now()

def export_excel(data_prest, data_prest_public):
    
  # Hoja 1 - Prestaciones activas sin PA

  wb = Workbook()
  ws = wb.active
  ws.title = "Prestaciones activas sin PA"

  headers_resumen = ["PRESTACION ID", "ALUMNO", "FEC. ACTIVACION", "FEC. DE ULTIMA BAJA", "DIAS SIN PA", 
                     "DIAGNOSTICO", "NIVEL", "TURNO", "COORDINADORA", "ESCUELA", "ESC. DIRECCION", 
                     "ESC. MAIL", "ESC. TEL 1", "ESC. TEL 2", "ESC. LOCALIDAD", "ESC. PARTIDO"]
  
  ws.append(headers_resumen)

  for cell in ws[1]:
      cell.font = Font(bold=True)

  for row in data_prest:
      ws.append(row)

  # Segunda hoja (Prest. sin pa para publicar)
  ws2 = wb.create_sheet(title="Prest. sin PA para publicar")

  headers_public = ["PRESTACION ID", "NIVEL", "TURNO", "ESCUELA", "LOCALIDAD", "PARTIDO",
                    "COORDINADORA", "MAIL COORDINADORA"]
  
  ws2.append(headers_public)

  for cell in ws2[1]:
    cell.font = Font(bold=True)

  for row in data_prest_public:
    ws2.append(row)

  nombre_archivo = f"reporte_comunicacion_{today.strftime('%Y-%m-%d')}.xlsx"
  wb.save(nombre_archivo)
  print(f"Archivo Excel generado: {nombre_archivo}")
  return nombre_archivo

def enviar_correo(nombre_archivo):
  try:
    yag = yagmail.SMTP(MAIL_AUTOR, APP_GMAIL_PASS)
    yag.send(
      to=MAIL_DESTINO,
      subject="Reporte general de Comunicación",
      contents= """Buenos días, se adjunta el reporte semanal del área de Comunicación.
              \nSaludos,\nMariano López - Ailes Inclusión.""",
      attachments=nombre_archivo
    )
    print("Correo enviado correctamente.")
  except Exception as e:
    print("Error al enviar el correo:", e)